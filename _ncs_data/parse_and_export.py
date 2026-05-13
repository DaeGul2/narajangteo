"""NCS 채용대행 수주 현황 v3 — GPT 정규화 매핑 기반 (공공기관/원본만 분리)."""
from __future__ import annotations
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
INST_MAP_PATH = ROOT / 'inst_map.json'

# 연도별 출처 블로그 URL
BLOG_URLS = {
    '2020': 'https://blog.naver.com/6860code/222251344144',
    '2021': 'https://blog.naver.com/6860code/222617714597',
    '2022': 'https://blog.naver.com/6860code/222976617950',
    '2023': 'https://blog.naver.com/6860code/223317003344',
    '2024': 'https://blog.naver.com/6860code/223725565087',
    '2025': 'https://blog.naver.com/6860code/224137651754',
}

# ─────────────────────────────────────────────────────────────
# 대행사 수동 매핑 (공백 차이)
# ─────────────────────────────────────────────────────────────
AGENCY_FIX: dict[str, str] = {
    '잡플러스 에이치알': '잡플러스에이치알',
    '에이디링크 연구소': '에이디링크연구소',
}

# 노이즈 패턴 (본문 텍스트가 잘못 잡힌 것)
NOISE_RES = [
    re.compile(r'^전자조달시스템'),
    re.compile(r'^※'),
    re.compile(r'^\*'),
    re.compile(r'채용공고\s*상의'),
    re.compile(r'^[^가-힯A-Za-z]+$'),
    re.compile(r'알려주세요'),
    re.compile(r'용역계약목록'),
    re.compile(r'^[0-9]+$'),
    # 블로그 footer/링크 노이즈
    re.compile(r'bit\.ly'),
    re.compile(r'https?://'),
    re.compile(r'📚'),
    re.compile(r'오카방|오픈채팅|오류사항|모듈의\s*코드|대행사\s*총정리|대행사\s*정보방'),
    re.compile(r'구분했습니다'),
    re.compile(r'시험일정'),
    # 단어 자체가 기관명 아닌 footer 잔류
    re.compile(r'^수의계약$'),
    re.compile(r'^나라장터'),
]

# footer 진입 시 그 이후 모든 라인 무시 (대행사 컨텍스트 종료)
FOOTER_TRIGGER_RE = re.compile(
    r'^※|나라장터\s*용역계약목록|bit\.ly|📚|모듈의\s*코드|대행사\s*총정리|대행사\s*정보방|오류사항|시험일정'
)


def is_noise(s: str) -> bool:
    if not s or len(s) < 2:
        return True
    if len(s) >= 60:  # 비정상적으로 긴 텍스트 = 본문 안내문 흘러들어옴
        return True
    return any(p.search(s) for p in NOISE_RES)


# ─────────────────────────────────────────────────────────────
# 연도 파일 파싱 → (year, agency, season, inst_full, date)
# ─────────────────────────────────────────────────────────────
def parse_year(year: str, path: Path) -> list[tuple]:
    lines = [l.strip() for l in path.read_text(encoding='utf-8').splitlines() if l.strip()]
    rows: list[tuple] = []
    cur_agency: str | None = None
    cur_season: str | None = None
    buffer: list[str] = []

    # 항목이 이런 패턴으로 시작하면 직전 기관명 prefix 가 빠진 것 (콤마로 잘못 쪼개진 것)
    PROJECT_PREFIX_RE = re.compile(
        r'^(?:'
        r'제?\s*\d+\s*(?:차|회|기|호선|분기|급|직급)|'
        r'\d+\s*[·∙\-,\s]+\s*\d+\s*직급|'
        r'\d+\s*급사원|'
        r'[가-힯]{1,4}\s*\d+\s*급|'   # 행정7급, 기술6급, 체육7급, 화학5급 등
        r'재공고|정정공고|수시채용|수시|'
        r'상반기|하반기|'
        r'보훈대상자|보훈|'
        r'경력전문가|경력직|경력|신입|고졸|'
        r'사회형평적인재|사회형평적|민간경력|'
        r'체험형\s*인턴|체험형|채용형인턴|청년인턴|인턴사원|인턴|'
        r'연구위원|제한경쟁|국가유공자|기록물전문요원|신규직원|청년인재|'
        r'무기계약직|기간제근로자|기간제|정규직|별정직|특정직|전문직|'
        r'사무직|일반직|기능원|기능직|기술직|연구직|시설직|관리직|운영직|'
        r'공무직|업무직원|업무직|업무지원직|전임직|지원직|방재직|통상직|학연지원직|시설서비스직|'
        r'공동채용|통합채용|운영부문|운영지원|행정직|'
        r'장애인|장애인계약직|'
        r'[가-힯]{2,5}직'   # 사무직/운영직/방재직 등 일반 ~직
        r')(?:[\s,(]|$)'
    )

    def get_inst_prefix(item: str) -> str:
        """item 텍스트에서 사업/회차 부분을 떼고 기관명만."""
        # 괄호 제거
        s = re.sub(r'\([^)]*\)', '', item).strip()
        # 사업 키워드 위치 찾기 (단어 시작)
        m = re.search(
            r'\s+(?:제?\s*\d+\s*(?:차|회|기|호선|분기|급|직급)|'
            r'\d+\s*급사원|행정\s*\d+\s*급|기술\s*\d+\s*급|'
            r'재공고|정정공고|수시(?:채용)?|'
            r'상반기|하반기|보훈대상자|보훈|'
            r'경력(?:전문가|직)?|신입|고졸|'
            r'사회형평적(?:인재)?|민간경력|체험형(?:\s*인턴)?|채용형인턴|청년인턴|인턴(?:사원)?|'
            r'무기계약직|기간제(?:근로자)?|정규직|별정직|특정직|전문직|사무직|일반직|'
            r'공무직|업무직(?:원)?|업무지원직|전임직|지원직|방재직|통상직|학연지원직|시설서비스직|'
            r'기능원|기능직|기술직|연구직|시설직|관리직|운영직|'
            r'공동채용|통합채용|운영부문|운영지원|행정직|장애인)',
            s,
        )
        if m:
            return s[:m.start()].strip()
        return s

    def flush():
        nonlocal buffer
        if not (buffer and cur_agency and cur_season):
            buffer = []
            return
        text = ' '.join(buffer)
        items, cur, depth = [], '', 0
        for ch in text:
            if ch == '(':
                depth += 1; cur += ch
            elif ch == ')':
                depth -= 1; cur += ch
            elif ch == ',' and depth == 0:
                if cur.strip(): items.append(cur.strip())
                cur = ''
            else:
                cur += ch
        if cur.strip():
            items.append(cur.strip())

        # ★ 사업 키워드로 시작하는 항목은 직전 항목의 기관명 prefix 와 결합
        # (작성자가 'X공단 제1회, 2회 재공고' 처럼 콤마로 같은 기관 다중 회차를 적은 케이스)
        fixed: list[str] = []
        last_inst_prefix = ''
        for it in items:
            it_s = it.strip().rstrip(',').strip()
            if not it_s: continue
            if PROJECT_PREFIX_RE.match(it_s) and last_inst_prefix:
                it_s = last_inst_prefix + ' ' + it_s
            fixed.append(it_s)
            last_inst_prefix = get_inst_prefix(it_s)

        for it in fixed:
            m = re.match(r'^(.+?)\(([^)]*)\)\s*$', it)
            if m:
                inst_full, date = m.group(1).strip(), m.group(2).strip()
            else:
                inst_full, date = it, ''
            inst_full = re.sub(r'\s+', ' ', inst_full).strip()
            if inst_full:
                rows.append((year, cur_agency, cur_season, inst_full, date))
        buffer = []

    for line in lines:
        # footer 진입 → 그 이후 라인 모두 무시 (대행사 컨텍스트 종료)
        if FOOTER_TRIGGER_RE.match(line):
            flush()
            cur_agency = None
            cur_season = None
            continue
        if line.startswith('*'):
            flush()
            cur_agency = line.lstrip('* ').strip()
            cur_season = None
        elif line in ('상반기', '하반기'):
            flush()
            cur_season = line
        else:
            if cur_agency:
                buffer.append(line)
    flush()
    return rows


# ─────────────────────────────────────────────────────────────
# 엑셀 스타일
# ─────────────────────────────────────────────────────────────
HEAD_FONT = Font(bold=True, color='FFFFFF', size=11)
HEAD_FILL = PatternFill('solid', fgColor='2563EB')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER


def main():
    # ─ 1. GPT 정규화 매핑 로드 ─
    if not INST_MAP_PATH.exists():
        print('❌ inst_map.json 없음 — 먼저 normalize_with_gpt.py 실행')
        return
    inst_map: dict[str, str] = json.loads(INST_MAP_PATH.read_text(encoding='utf-8'))
    print(f'GPT 매핑 로드: {len(inst_map)}개')

    # ─ 2. 파싱 ─
    raw_rows: list[tuple] = []
    counts = {}
    for year in ('2020', '2021', '2022', '2023', '2024', '2025'):
        rows = parse_year(year, ROOT / f'{year}.txt')
        counts[year] = len(rows)
        raw_rows.extend(rows)
    print(f'raw rows: {len(raw_rows)} {counts}')

    # ─ 3. 노이즈 필터 ─
    n_before = len(raw_rows)
    raw_rows = [r for r in raw_rows if not is_noise(r[3])]
    print(f'노이즈 제거: {n_before - len(raw_rows)}건 → {len(raw_rows)}')

    # ─ 4. 대행사 정규화 (공백 통합) ─
    all_agencies = [r[1] for r in raw_rows]
    # 매뉴얼 fix 적용 후 공백 정규화로 빈도 최대 표기 통일
    norm_groups: dict[str, Counter] = defaultdict(Counter)
    for n in all_agencies:
        fixed = AGENCY_FIX.get(n, n)
        norm = re.sub(r'[\s·\-_]+', '', fixed)
        norm_groups[norm][fixed] += 1
    canon_agency: dict[str, str] = {}
    for norm, ct in norm_groups.items():
        best = ct.most_common(1)[0][0]
        for variant in ct:
            canon_agency[variant] = best
    agency_full_map = {orig: canon_agency.get(AGENCY_FIX.get(orig, orig), orig) for orig in all_agencies}

    # ─ 5. 최종 행 생성: (year, agency, season, inst_norm, date, inst_original, blog_url) ─
    final_rows: list[tuple] = []
    not_in_map = 0
    for year, agency, season, inst_full, date in raw_rows:
        agency_final = agency_full_map.get(agency, agency)
        if inst_full in inst_map:
            inst_norm = inst_map[inst_full]
        else:
            inst_norm = inst_full
            not_in_map += 1
        final_rows.append((year, agency_final, season, inst_norm, date, inst_full, BLOG_URLS.get(year, '')))
    if not_in_map:
        print(f'⚠️ GPT 매핑에 없는 표기: {not_in_map}건 (원본 그대로 사용)')

    # ─ 6. 통계 ─
    ag_orig = len(set(all_agencies))
    ag_after = len(set(r[1] for r in final_rows))
    in_orig = len(set(r[3] for r in raw_rows))
    in_after = len(set(r[3] for r in final_rows))
    print(f'대행사 unique: {ag_orig} → {ag_after}')
    print(f'기관 unique: {in_orig} → {in_after}')

    # ─ 7. 엑셀 ─
    wb = Workbook()
    ws = wb.active
    ws.title = '전체'
    headers = ['연도', '채용대행사', '시기', '공공기관', '계약일', '원본', '출처 블로그']
    ws.append(headers)
    style_header(ws, len(headers))
    for r in final_rows:
        ws.append(list(r))
    for i, w in enumerate([8, 22, 8, 36, 22, 60, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    # 출처 컬럼은 하이퍼링크
    from openpyxl.styles import Font as _F
    link_font = _F(color='0563C1', underline='single', size=11)
    for row in range(2, ws.max_row + 1):
        c = ws.cell(row=row, column=7)
        if c.value:
            c.hyperlink = c.value
            c.font = link_font

    years = ['2020', '2021', '2022', '2023', '2024', '2025']

    # 피벗: 대행사 × 연도
    ws3 = wb.create_sheet('피벗_대행사X연도')
    pivot: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in final_rows:
        pivot[r[1]][r[0]] += 1
    ws3.append(['채용대행사'] + years + ['합계'])
    style_header(ws3, len(years) + 2)
    plist = [(ag, [pivot[ag].get(y, 0) for y in years]) for ag in pivot]
    plist.sort(key=lambda x: -sum(x[1]))
    for ag, cnt in plist:
        ws3.append([ag] + cnt + [sum(cnt)])
    ws3.column_dimensions['A'].width = 28
    for c in range(2, 9):
        ws3.column_dimensions[get_column_letter(c)].width = 9
    ws3.freeze_panes = 'B2'

    # 피벗: 기관 × 연도 (대행사 누구가 했는지)
    inst_year_ag: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for r in final_rows:
        if r[3]:
            inst_year_ag[r[3]][r[0]].append(r[1])
    ws4 = wb.create_sheet('피벗_기관X연도_대행사')
    ws4.append(['공공기관'] + years + ['총 수주'])
    style_header(ws4, len(years) + 2)
    rows4 = []
    for inst, yd in inst_year_ag.items():
        total = sum(len(yd.get(y, [])) for y in years)
        row = [inst]
        for y in years:
            ags = yd.get(y, [])
            row.append(', '.join(dict.fromkeys(ags)))
        row.append(total)
        rows4.append((row, total))
    rows4.sort(key=lambda x: -x[1])
    for row, _ in rows4:
        ws4.append(row)
    ws4.column_dimensions['A'].width = 32
    for c in range(2, 8):
        ws4.column_dimensions[get_column_letter(c)].width = 26
    ws4.column_dimensions['H'].width = 10
    ws4.freeze_panes = 'B2'

    # 피벗: 기관별 총 수주 + 연도별 합
    ws_inst = wb.create_sheet('기관_총수주순위')
    ws_inst.append(['공공기관', '총 수주'] + years)
    style_header(ws_inst, 8)
    inst_total: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in final_rows:
        if r[3]:
            inst_total[r[3]][r[0]] += 1
    inst_sorted = sorted(inst_total.items(), key=lambda x: -sum(x[1].values()))
    for inst, yc in inst_sorted:
        ws_inst.append([inst, sum(yc.values())] + [yc.get(y, 0) for y in years])
    ws_inst.column_dimensions['A'].width = 36
    ws_inst.column_dimensions['B'].width = 10
    for c in range(3, 9):
        ws_inst.column_dimensions[get_column_letter(c)].width = 8
    ws_inst.freeze_panes = 'C2'
    ws_inst.auto_filter.ref = ws_inst.dimensions

    # 연도별 대행사 카운트
    for year in years:
        agg = Counter(r[1] for r in final_rows if r[0] == year)
        ws2 = wb.create_sheet(f'{year}_대행사별')
        ws2.append(['채용대행사', '수주 건수'])
        style_header(ws2, 2)
        for ag, cnt in sorted(agg.items(), key=lambda x: -x[1]):
            ws2.append([ag, cnt])
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 12
        ws2.freeze_panes = 'A2'

    # 인사바른 전용
    ws5 = wb.create_sheet('인사바른_수주현황')
    ws5.append(['연도', '시기', '공공기관', '계약일', '원본', '출처 블로그'])
    style_header(ws5, 6)
    for r in final_rows:
        if '인사바른' in r[1]:
            ws5.append([r[0], r[2], r[3], r[4], r[5], r[6]])
    for i, w in enumerate([8, 8, 36, 22, 60, 50], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = 'A2'
    ws5.auto_filter.ref = ws5.dimensions
    for row in range(2, ws5.max_row + 1):
        c = ws5.cell(row=row, column=6)
        if c.value:
            c.hyperlink = c.value
            c.font = link_font

    # GPT 정규화 로그
    ws6 = wb.create_sheet('정규화_매핑로그')
    ws6.append(['원본 표기', '정제 결과', '발생 빈도'])
    style_header(ws6, 3)
    raw_count = Counter(r[3] for r in raw_rows)
    log_rows = []
    for orig, canon in inst_map.items():
        if orig != canon and orig in raw_count:
            log_rows.append((orig, canon, raw_count[orig]))
    log_rows.sort(key=lambda x: -x[2])
    for r in log_rows:
        ws6.append(list(r))
    ws6.column_dimensions['A'].width = 42
    ws6.column_dimensions['B'].width = 36
    ws6.column_dimensions['C'].width = 10
    ws6.freeze_panes = 'A2'
    ws6.auto_filter.ref = ws6.dimensions

    out = ROOT.parent / 'NCS_채용대행_연도별수주현황.xlsx'
    wb.save(out)
    print(f'\n저장: {out}')
    print(f'정규화 로그(원본≠정제): {len(log_rows)}건')


if __name__ == '__main__':
    main()
