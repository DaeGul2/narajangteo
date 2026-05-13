"""엑셀 → agency_history.json — 백엔드가 이전 채용대행사 빠르게 검색하기 위한 인덱스."""
from __future__ import annotations
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT.parent / 'NCS_채용대행_연도별수주현황.xlsx'
OUT = ROOT / 'agency_history.json'

wb = load_workbook(XLSX, data_only=True)
ws = wb['전체']
idx: dict[str, list[dict]] = defaultdict(list)

for row in ws.iter_rows(min_row=2, values_only=True):
    year, agency, season, inst, date, full, url = row
    if not (inst and agency):
        continue
    idx[inst].append({
        'year': str(year),
        'agency': agency,
        'season': season,
        'date': date,
    })

# year 내림차순 정렬
for k in idx:
    idx[k].sort(key=lambda r: (r['year'], r['season'] or ''), reverse=True)

OUT.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'기관 {len(idx)}개, 총 {sum(len(v) for v in idx.values())}건 인덱싱')
print(f'저장: {OUT}')
